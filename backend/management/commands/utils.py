import openpyxl
from openpyxl.styles import Alignment
from geopy import Nominatim
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

from backend.models import Product


async def get_location(message):
    geolocator = Nominatim(user_agent='geoapiExercises')
    location = geolocator.reverse(
        str(message.location.latitude) + ',' + str(message.location.longitude)
    )

    # Extract the address information from the location object
    address = location.raw['address']
    city = address.get('city', '')
    state = address.get('state', '')
    country = address.get('country', '')
    postcode = address.get('postcode', '')

    # Construct the full address string
    address_parts = [city, state, country, postcode]
    full_address = ', '.join([part for part in address_parts if part])

    # Create a dictionary with the location information
    location_info = {
        'latitude': location.latitude,
        'longitude': location.longitude,
        'address': full_address,
        'city': city,
        'state': state,
        'country': country,
        'postcode': postcode,
    }
    return location_info


async def write_to_excel(location, user, cart):
    try:
        workbook = openpyxl.load_workbook('purchases.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active

    try:
        # Find the next available row in the sheet
        row = sheet.max_row + 1

        # Write the location, user, and shopping cart information to the sheet
        sheet.cell(row=row, column=1, value=user.username)
        sheet.cell(row=row, column=2, value=str(location))
        for item in cart:
            sheet.cell(row=row, column=3, value=item.product.name)
            sheet.cell(row=row, column=4, value=item.amount)
            row += 1

        # Adjust the column widths and cell alignment
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 10
        for row in sheet.iter_rows(min_row=1, max_row=row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)

        # Save the Excel file
        workbook.save('purchases.xlsx')
        return True
    except Exception:
        return False


async def create_keyboard(message, subcategory, category):
    products_markup = InlineKeyboardMarkup()
    async for product in Product.objects.filter(subcategory=subcategory):
        button = InlineKeyboardButton(text=product.name, callback_data=f"product_{product.name}")
        products_markup.add(button)
    back_button = InlineKeyboardButton(
        text="Вернуться к подкатегориям",
        callback_data=f"category_{category}_"
                      f"{subcategory}_subcategory_"
                      f"go_sub"
    )
    products_markup.add(back_button)
    await message.answer('Товары:', reply_markup=products_markup)
